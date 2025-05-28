import argparse
import csv
import os
from pathlib import Path
import shutil
import sys
import logging
from tqdm import tqdm
from openpyxl import load_workbook
import pandas as pd

__author__ = "Miguel Angel Salinas Gancedo"
__copyright__ = "Simur"
__license__ = "MIT"

_logger = logging.getLogger(__name__)

files_to_check = []

def parse_extensions(ext_string):
    # Split by comma, strip whitespace, ensure each starts with dot, and make a set
    return {ext.strip() if ext.strip().startswith('.') else '.' + ext.strip()
            for ext in ext_string.split(',')}

def parse_args(args):
    """Parse command line parameters

    Args:
      args (List[str]): command line parameters as list of strings
          (for example  ``["--help"]``).

    Returns:
      :obj:`argparse.Namespace`: command line parameters namespace
    """
    parser = argparse.ArgumentParser(description="BIN to CSV Converter")

    parser.add_argument(
        "-sr",
        "--source-root",
        required=True,
        dest="source_root",
        help="Source root folder",
    )
    parser.add_argument(
        "-df",
        "--destination-file",
        required=True,
        dest="destination_file",
        help="Destination file",
    )                 
    parser.add_argument(
        "-v",
        "--verbose",
        dest="loglevel",
        help="set loglevel to INFO",
        action="store_const",
        const=logging.INFO,
    )
    parser.add_argument(
        "-vv",
        "--very-verbose",
        dest="loglevel",
        help="set loglevel to DEBUG",
        action="store_const",
        const=logging.DEBUG,
    )

    return parser.parse_args(args)

def setup_logging(loglevel):
    """Setup basic logging

    Args:
      loglevel (int): minimum loglevel for emitting messages
    """
    logformat = "[%(asctime)s] %(levelname)s:%(name)s:%(message)s"
    logging.basicConfig(
        level=loglevel, stream=sys.stdout, format=logformat, datefmt="%Y-%m-%d %H:%M:%S"
    )

def check_activity_register(file):
    # Load the workbook
    wb = load_workbook(file)

    # Select a sheet (by name or active)
    sheet = wb.active

    # Check if exist final date
    target = 'Devolución acelerómetro de muñeca'
    found_cell = None

    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == target:
                if (row[cell.column + 2].value is not None):
                    found_cell = cell
                    break
        if found_cell:
            break

    if found_cell:
        print(f"Found '{target}' at cell {found_cell.coordinate}")
        return True, row[cell.column + 2].value
    else:
        return False, None    

def get_files(args):
    for root, dirs, files in os.walk(args.source_root):
        for file in files:
            # get participant name           
            _, ext = os.path.splitext(file)         

            # get only activity registers
            if ext == ".BIN":
                files_to_check.append((root, ext, file, None, None))
            elif (ext == ".xlsx" and "RegistroActividades" in file):
                try:
                    exist_final_date = check_activity_register(os.path.join(root, file))
                except:
                    _logger.error("Error checking this activity register: " + file)                    
            
                files_to_check.append((root, ext, file, exist_final_date[0], exist_final_date[1]))
      
    files_updated = [(str(Path(root).relative_to(args.source_root)), ext, file, exist_final_date, final_date_value) for root, ext, file, exist_final_date, final_date_value in files_to_check]

    files_ordered = sorted(files_updated)

    return files_ordered

def create_csv(args, files_to_check):
    with open(args.destination_file, mode='w', newline='') as file:
        writer = csv.writer(file)
        writer.writerows(files_to_check)

        h = ["PARTICIPANT", "EXTENSION", "FILE", "EXIST_FINAL_DATE", "FINAL_DATE_VALUE"]
        df = pd.read_csv(args.destination_file, header=None, names=h)

        df.to_excel(args.destination_file.replace("csv", "xlsx"), sheet_name="Resume", index=False)       

_logger.info("Starting checker ...")

args = parse_args(sys.argv[1:])
setup_logging(args.loglevel)

_logger.info("Get files to be checked ...")
files_to_check = get_files(args)

_logger.info("Create CSV report ...")
create_csv(args, files_to_check)

_logger.info("Ending checker ...")